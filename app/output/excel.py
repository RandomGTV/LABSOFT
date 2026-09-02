"""Excel and CSV import/export.

Import previews what it will do and reports what it skipped before writing
anything, so a badly-shaped sheet cannot quietly half-load the lab's test list.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from ..core import billing, ranges as rng
from ..db import queries as q


# ------------------------------------------------------------------- export

def write_sheet(path: Path, headers: Sequence[str], rows: Iterable[Sequence],
                title: str = "Sheet1") -> Path:
    """Write .xlsx when openpyxl is available, otherwise .csv."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            path = path.with_suffix(".csv")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = title[:31]
            ws.append(list(headers))
            head_fill = PatternFill("solid", fgColor="0F5C73")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = head_fill
                cell.alignment = Alignment(vertical="center")
            for r in rows:
                ws.append(list(r))
            for i, h in enumerate(headers, start=1):
                longest = max([len(str(h))] +
                              [len(str(ws.cell(row=r, column=i).value or ""))
                               for r in range(2, ws.max_row + 1)] or [10])
                ws.column_dimensions[get_column_letter(i)].width = min(46, longest + 3)
            ws.freeze_panes = "A2"
            wb.save(path)
            return path

    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(list(headers))
        writer.writerows([list(r) for r in rows])
    return path


def export_ledger(path: Path, rows: Sequence[dict]) -> Path:
    headers = ["Report No", "Date", "Patient", "Referred by", "Charged",
               "Discount", "Net", "Paid", "Balance", "Commission"]
    body = []
    for r in rows:
        body.append([
            r["report_no"],
            (q.to_dt(r["received_at"]) or datetime.now()).strftime("%d-%m-%Y"),
            r["patient_name"], r.get("referrer_name") or "",
            billing.to_rupees(r["gross_paise"]),
            billing.to_rupees(r["discount_paise"]),
            billing.to_rupees(r["net_paise"]),
            billing.to_rupees(r["paid_paise"]),
            billing.to_rupees(r["balance_paise"]),
            billing.to_rupees(r["commission_paise"]),
        ])
    return write_sheet(path, headers, body, "Ledger")


def export_tests(path: Path) -> Path:
    headers = ["Code", "Name", "Group", "Unit", "Decimals", "Type", "Options",
               "Formula", "Rate", "TAT hours", "Normal Value"]
    body = []
    for t in q.list_tests(include_inactive=True):
        ranges = q.ranges_for_test(t["id"])
        normal = ranges[0]["display_text"] if ranges else ""
        body.append([t["code"], t["name"], t["group_name"], t["unit"], t["decimals"],
                     t["result_type"], t["options"], t["formula"],
                     billing.to_rupees(t["rate_paise"]), t["tat_hours"], normal])
    return write_sheet(path, headers, body, "Tests")


# ------------------------------------------------------------------- import

_ALIASES = {
    "code": "code", "test code": "code", "short code": "code",
    "name": "name", "test": "name", "test name": "name", "description": "name",
    "test description": "name", "investigation": "name",
    "group": "group_name", "group name": "group_name", "department": "group_name",
    "category": "group_name", "section": "group_name",
    "unit": "unit", "units": "unit",
    "formula": "formula", "calculation": "formula",
    "rate": "rate", "price": "rate", "amount": "rate", "charge": "rate",
    "tat": "tat", "tat hours": "tat", "turnaround": "tat",
    "normal": "normal", "normal value": "normal", "reference": "normal",
    "reference range": "normal", "range": "normal", "normal range": "normal",
    "decimals": "decimals", "type": "result_type", "options": "options",
}


def read_rows(path: Path) -> Tuple[List[str], List[List[str]]]:
    """Read a sheet as headers plus rows of strings."""
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        # Closed explicitly: a read-only workbook keeps the file open, and on
        # Windows the operator then cannot re-save or delete their own
        # spreadsheet until LabSoft is shut down.
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            data = [["" if c is None else str(c).strip() for c in row]
                    for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    else:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            data = [[(c or "").strip() for c in row] for row in csv.reader(fh, dialect)]

    data = [r for r in data if any(c for c in r)]
    if not data:
        return [], []
    return data[0], data[1:]


def map_columns(headers: Sequence[str]) -> Dict[str, int]:
    """Match the sheet's own column names to the fields we need."""
    found: Dict[str, int] = {}
    for i, h in enumerate(headers):
        key = _ALIASES.get((h or "").strip().lower())
        if key and key not in found:
            found[key] = i
    return found


def preview_tests_import(path: Path) -> dict:
    """Work out what an import would do, without changing anything."""
    headers, rows = read_rows(path)
    cols = map_columns(headers)
    if "name" not in cols:
        return {"ok": False, "reason":
                "No column of test names was found. The sheet needs a column "
                "headed Name, Test, or Test Description.",
                "headers": headers}

    existing = {t["code"].upper(): t for t in q.list_tests(include_inactive=True)}
    existing_names = {t["name"].strip().lower() for t in existing.values()}

    new, update, skipped = [], [], []
    seen_codes = set()

    for line, r in enumerate(rows, start=2):
        def cell(field: str) -> str:
            i = cols.get(field)
            return (r[i].strip() if i is not None and i < len(r) else "")

        name = cell("name")
        if not name:
            skipped.append((line, "no test name"))
            continue

        supplied = cell("code").upper()
        if supplied:
            code = supplied
        else:
            # Generated codes are checked against the database as well as the
            # sheet. Checking only the sheet let a made-up code collide with a
            # real test and overwrite it in place, silently rewriting every past
            # report that used it.
            code = _code_from_name(name, seen_codes | set(existing))

        if code in seen_codes:
            skipped.append((line, f"code {code} appears twice in the sheet"))
            continue
        seen_codes.add(code)

        record = {
            "code": code,
            "name": name,
            "group_name": cell("group_name") or "OTHER INVESTIGATIONS",
            "unit": cell("unit"),
            "formula": cell("formula"),
            "rate": cell("rate"),
            "tat": cell("tat"),
            "normal": cell("normal"),
            "line": line,
        }
        if code in existing:
            # Name what will be overwritten, so "1 test will be updated" is not
            # the only warning before an existing test is changed.
            record["replaces"] = existing[code]["name"]
            update.append(record)
        elif name.strip().lower() in existing_names:
            skipped.append((line, f"'{name}' already exists under another code"))
        else:
            new.append(record)

    return {"ok": True, "headers": headers, "columns": sorted(cols),
            "new": new, "update": update, "skipped": skipped}


def _code_from_name(name: str, taken: set) -> str:
    base = "".join(ch for ch in name.upper() if ch.isalnum())[:8] or "TEST"
    code = base
    n = 2
    while code in taken:
        code = f"{base[:6]}{n}"
        n += 1
    return code


def apply_tests_import(preview: dict, update_existing: bool = True) -> dict:
    """Write the rows the preview identified. Returns counts."""
    if not preview.get("ok"):
        return {"added": 0, "updated": 0, "failed": []}

    existing = {t["code"].upper(): t for t in q.list_tests(include_inactive=True)}
    added = updated = 0
    failed: List[Tuple[int, str]] = []

    batches = [("new", preview.get("new", []))]
    if update_existing:
        batches.append(("update", preview.get("update", [])))

    for kind, records in batches:
        for rec in records:
            try:
                payload = {
                    "code": rec["code"],
                    "name": rec["name"],
                    "group_name": rec["group_name"],
                    "unit": rec["unit"],
                    "decimals": 1,
                    "result_type": "numeric",
                    "options": "",
                    "formula": rec["formula"],
                    "rate_paise": billing.to_paise(rec["rate"]),
                    "tat_hours": _as_float(rec["tat"], 24.0),
                    "sort_order": 0,
                    "active": 1,
                }
                if kind == "update":
                    old = existing[rec["code"]]
                    payload["id"] = old["id"]
                    payload["decimals"] = old["decimals"]
                    payload["result_type"] = old["result_type"]
                    payload["options"] = old["options"]
                    # A BLANK CELL MEANS "LEAVE IT ALONE", not "set it to
                    # nothing". Only `formula` was guarded, so importing a
                    # sheet of two columns -- Code and Test, to rename a few
                    # tests -- rewrote the rate of every test it touched to
                    # ₹0.00, moved the turnaround to 24 hours, emptied the
                    # unit and dropped the department. Every bill after that
                    # charged nothing, and the dialog said only "1 existing
                    # test will be updated".
                    if not payload["formula"]:
                        payload["formula"] = old["formula"]
                    if not str(rec["rate"]).strip():
                        payload["rate_paise"] = old["rate_paise"]
                    if not str(rec["tat"]).strip():
                        payload["tat_hours"] = old["tat_hours"]
                    if not str(rec["unit"]).strip():
                        payload["unit"] = old["unit"]
                    if not str(rec["group_name"]).strip():
                        payload["group_name"] = old["group_name"]
                tid = q.save_test(payload)

                if rec["normal"]:
                    q.replace_ranges(tid, [_range_from_text(rec["normal"])])
                added += kind == "new"
                updated += kind == "update"
            except Exception as exc:
                failed.append((rec.get("line", 0), f"{rec['name']}: {exc}"))

    q.log_action("tests_imported", "tests", None,
                 f"{added} added, {updated} updated, {len(failed)} failed")
    return {"added": added, "updated": updated, "failed": failed}


def _as_float(text, default: float) -> float:
    try:
        return float(str(text).strip())
    except (TypeError, ValueError):
        return default


def _range_from_text(text: str) -> dict:
    """Turn '70 - 110mg/dl' or '< 200' into a rule, keeping the original wording
    for printing so the report reads exactly as the lab wrote it."""
    raw = (text or "").strip()
    body = raw.replace("–", "-").replace("—", "-")

    def num(s: str) -> Optional[float]:
        digits = ""
        for ch in s:
            if ch.isdigit() or ch in ".-" and (not digits or digits[-1].isdigit() is False):
                digits += ch
            elif ch.isdigit() or ch == ".":
                digits += ch
        try:
            return float(digits)
        except ValueError:
            return None

    import re
    m = re.search(r"(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)", body)
    if m:
        return {"rule_type": rng.RULE_RANGE, "low": float(m.group(1)),
                "high": float(m.group(2)), "display_text": raw, "sex": "any"}
    m = re.search(r"[<≤]\s*(-?\d+(?:\.\d+)?)", body)
    if m:
        return {"rule_type": rng.RULE_MAX, "high": float(m.group(1)),
                "display_text": raw, "sex": "any"}
    m = re.search(r"[>≥]\s*(-?\d+(?:\.\d+)?)", body)
    if m:
        return {"rule_type": rng.RULE_MIN, "low": float(m.group(1)),
                "display_text": raw, "sex": "any"}
    return {"rule_type": rng.RULE_TEXT, "text_value": raw, "display_text": raw,
            "sex": "any"}
